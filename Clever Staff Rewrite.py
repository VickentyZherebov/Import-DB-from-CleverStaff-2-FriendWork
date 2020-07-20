from typing import Dict

from openpyxl import load_workbook, Workbook

# загружаем эксель файл который надо пересобрать
from openpyxl.worksheet.worksheet import Worksheet

from Action import Action
from Candidate import Candidate

# Не понимаю, что тут происходит - 12 и 13 строка. Что значит "->"? Просто затупил
def load_candidates(workbook: Workbook) -> Dict[str, Candidate]:
    candidates_sheet: Worksheet = workbook['кандидаты']

    # Узнаем количество заполненных строк на странице "кандидаты",
    # вычетаем строку с заголовком и получаем количество кандидатов
    print(f'Количество кандидатов {candidates_sheet.max_row - 1}')

#  Создаем словарь "candidates" где ключом будет значение переменной в диапазоне от 2 до максимального количества
#     строк, а значением будет переменная candidates[local_id] в которую мы засовываем класс Candidate значения которого
#     заполняются значение при переборе значение переменной row_number, в конце мы делаем return candidates - то есть
#     записываем ключ-значение в словарь и бежим по новой до предела
    candidates = {}
    for row_number in range(2, candidates_sheet.max_row + 1):
        local_id = candidates_sheet[f'AB{row_number}'].value
        candidates[local_id] = Candidate(
            number=candidates_sheet[f'A{row_number}'].value,
            first_name=candidates_sheet[f'B{row_number}'].value,
            patronymic=candidates_sheet[f'C{row_number}'].value,
            last_name=candidates_sheet[f'D{row_number}'].value,
            desired_position=candidates_sheet[f'E{row_number}'].value,
            current_position=candidates_sheet[f'F{row_number}'].value,
            current_place=candidates_sheet[f'G{row_number}'].value,
            birth_date=candidates_sheet[f'H{row_number}'].value,
            sex=candidates_sheet[f'I{row_number}'].value,
            status=candidates_sheet[f'J{row_number}'].value,
            phone=candidates_sheet[f'K{row_number}'].value,
            email=candidates_sheet[f'L{row_number}'].value,
            skype=candidates_sheet[f'M{row_number}'].value,
            facebook=candidates_sheet[f'N{row_number}'].value,
            linkedin=candidates_sheet[f'O{row_number}'].value,
            type_of_employment=candidates_sheet[f'P{row_number}'].value,
            field_of_activity=candidates_sheet[f'Q{row_number}'].value,
            work_experience=candidates_sheet[f'R{row_number}'].value,
            salary=candidates_sheet[f'S{row_number}'].value,
            currency=candidates_sheet[f'T{row_number}'].value,
            language=candidates_sheet[f'U{row_number}'].value,
            region=candidates_sheet[f'V{row_number}'].value,
            date_of_adding=candidates_sheet[f'Z{row_number}'].value,
            local_id=local_id
        )

    return candidates


def load_history(workbook: Workbook, candidates: Dict[str, Candidate]): # так же не понятно, что тут происходит
    history_sheet: Worksheet = workbook['история']

    current_candidate = None # зачем None?
    for row_number in range(3, history_sheet.max_row): # здесь понятно, что происходит так же как и с кандидатами -
        # 24 строка
        local_id = history_sheet[f'D{row_number}'].value

        if local_id: # тут условие мне тоже не ясно. Не понятна фраза "if local_id:" - если local_id что?)
            current_candidate = candidates[local_id]

        if current_candidate:
            current_candidate.actions.append(Action(
                when=history_sheet[f'E{row_number}'].value,
                who=history_sheet[f'F{row_number}'].value,
                action=history_sheet[f'G{row_number}'].value
            )) # непонятно, как должно происходить смещение вниз по пустым полям


workbook = load_workbook('AIHUB.xlsx')
candidates = load_candidates(workbook)
load_history(workbook, candidates)

# Выдает ошибку
# for local_id, candidate in candidates.items()[:20]:
#     Traceback(most
#     recent
#     call
#     last):
#     File
#     "C:/PyProgects/Clever Staff Rewrite.py", line
#     79, in < module >
#     for local_id, candidate in candidates.items()[:20]:
# TypeError: 'dict_items'
# object is not subscriptable
    print(candidate)
