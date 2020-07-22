from typing import Dict

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from Action import Action
from Candidate import Candidate
from Comment import Comment


def load_candidates(workbook: Workbook) -> Dict[str, Candidate]:
    candidates_sheet: Worksheet = workbook['кандидаты']

    candidates: Dict[str, Candidate] = {}
    for row_number in range(2, candidates_sheet.max_row + 1):
        local_id: str = candidates_sheet[f'AB{row_number}'].value
        candidates[local_id] = Candidate(
            number=candidates_sheet[f'A{row_number}'].value,
            first_name=candidates_sheet[f'B{row_number}'].value,
            patronymic=candidates_sheet[f'C{row_number}'].value,
            last_name=candidates_sheet[f'D{row_number}'].value,
            desired_position=candidates_sheet[f'E{row_number}'].value,
            current_position=candidates_sheet[f'F{row_number}'].value,
            current_place=candidates_sheet[f'G{row_number}'].value,
            birth_date=candidates_sheet[f'H{row_number}'].value,
            sex=candidates_sheet[f'I{row_number}'].value,
            status=candidates_sheet[f'J{row_number}'].value,
            phone=candidates_sheet[f'K{row_number}'].value,
            email=candidates_sheet[f'L{row_number}'].value,
            skype=candidates_sheet[f'M{row_number}'].value,
            facebook=candidates_sheet[f'N{row_number}'].value,
            linkedin=candidates_sheet[f'O{row_number}'].value,
            type_of_employment=candidates_sheet[f'P{row_number}'].value,
            field_of_activity=candidates_sheet[f'Q{row_number}'].value,
            work_experience=candidates_sheet[f'R{row_number}'].value,
            salary=candidates_sheet[f'S{row_number}'].value,
            currency=candidates_sheet[f'T{row_number}'].value,
            language=candidates_sheet[f'U{row_number}'].value,
            region=candidates_sheet[f'V{row_number}'].value,
            date_of_adding=candidates_sheet[f'Z{row_number}'].value,
            local_id=local_id
        )

    return candidates


def load_history(workbook: Workbook, candidates: Dict[str, Candidate]):
    history_sheet: Worksheet = workbook['история']

    current_candidate: [Candidate] = None
    for row_number in range(3, history_sheet.max_row + 1):
        local_id = history_sheet[f'D{row_number}'].value

        if local_id is not None:
            current_candidate = candidates[local_id]

        if current_candidate is not None:
            current_candidate.actions.append(Action(
                when=history_sheet[f'E{row_number}'].value,
                who=history_sheet[f'F{row_number}'].value,
                action=history_sheet[f'G{row_number}'].value
            ))


def load_comments(workbook: Workbook, candidates: Dict[str, Candidate]):
    comment_sheet: Worksheet = workbook['комментарии']

    current_candidate: [Candidate] = None
    for row_number in range(2, comment_sheet.max_row + 1):
        local_id = comment_sheet[f'D{row_number}'].value

        if local_id is not None:
            current_candidate = candidates[local_id]

        if current_candidate is not None:
            current_candidate.comments.append(Comment(
                when=comment_sheet[f'E{row_number}'].value,
                who=comment_sheet[f'F{row_number}'].value,
                text=comment_sheet[f'G{row_number}'].value
            ))
