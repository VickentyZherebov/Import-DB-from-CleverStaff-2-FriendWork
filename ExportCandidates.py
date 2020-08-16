from typing import Dict, Callable
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from Candidate import Candidate


class ExportColumn:
    def __init__(self, name: str, get_value: Callable[[Candidate], str]):
        self.name = name
        self.get_value = get_value


_columns = [
    ExportColumn("Вакансия", lambda candidate: "Название вакансии"),
    ExportColumn("Ответственный за вакансию", lambda candidate: "Ответственный за вакансию"),
    ExportColumn("Создатель вакансии", lambda candidate: "Создатель вакансии"),
    ExportColumn("Дата создания вакансии", lambda candidate: "Дата создания вакансии"),
    ExportColumn("Заказчик", lambda candidate: "Заказчик"),
    ExportColumn("Источник", lambda candidate: candidate.linkedin_is_link),
    ExportColumn("Ссылка на резюме", lambda candidate: candidate.linkedin),
    ExportColumn("Способ занесения", lambda candidate: candidate.status),
    ExportColumn("Дата создания кандидата", lambda candidate: candidate.date_of_adding),
    ExportColumn("Создатель кандидата", lambda candidate: "Создатель кандидата"),
    ExportColumn("Ответственный за кандидата", lambda candidate: "Ответственный за кандидата"),
    ExportColumn("Фамилия", lambda candidate: candidate.last_name),
    ExportColumn("Имя", lambda candidate: candidate.first_name),
    ExportColumn("Отчество", lambda candidate: candidate.patronymic),
    ExportColumn("Регион", lambda candidate: candidate.region),
    ExportColumn("Возраст", lambda candidate: ""),
    ExportColumn("Дата Рождения", lambda candidate: candidate.birth_date),
    ExportColumn("Email", lambda candidate: candidate.email),
    ExportColumn("Skype", lambda candidate: candidate.skype),
    ExportColumn("Контакты", lambda candidate: ""),
    ExportColumn("Телефон", lambda candidate: candidate.phone),
    ExportColumn("Зарплата", lambda candidate: candidate.salary),
    ExportColumn("Комментарий - Email", lambda candidate: candidate.email_comment),
    ExportColumn("Комментарий - Phone", lambda candidate: candidate.phone_comment),
    ExportColumn("Комментарий - Linkedin", lambda candidate: candidate.linkedin_comment),
]


def export_candidates(workbook: Workbook, candidates: Dict[str, Candidate]):
    worksheet: Worksheet = workbook.create_sheet("результат")
    max_comments = count_max_comments(candidates)

    columns_count = len(_columns)
    for column_index in range(0, columns_count):
        worksheet.cell(
            row=1,
            column=column_index + 1,
            value=_columns[column_index].name
        )

    for column_index in range(0, max_comments):
        worksheet.cell(
            row=1,
            column=columns_count + column_index + 1,
            value="Комментарий"
        )

    out_row = 2
    for local_id, candidate in candidates.items():
        for column_index in range(0, columns_count):
            worksheet.cell(
                row=out_row,
                column=column_index + 1,
                value=_columns[column_index].get_value(candidate)
            )
        comment_index = 0
        for comment in candidate.comments:
            worksheet.cell(
                row=out_row,
                column=columns_count + comment_index + 1,
                value=f'{comment.when}\n{comment.who}\n{comment.text}'
            )
            comment_index = comment_index + 1
        for action in candidate.actions:
            worksheet.cell(
                row=out_row,
                column=columns_count + comment_index + 1,
                value=f'{action.when}\n{action.who}\n{action.action}'
            )
            comment_index = comment_index + 1
        out_row = out_row + 1


def count_max_comments(candidates: Dict[str, Candidate]) -> int:
    max_comments = 0
    for candidate in candidates.values():
        candidate_comments = len(candidate.comments) + len(candidate.actions)
        if max_comments < candidate_comments:
            max_comments = candidate_comments
    return max_comments
