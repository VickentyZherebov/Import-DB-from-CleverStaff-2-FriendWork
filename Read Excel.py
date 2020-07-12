from openpyxl import load_workbook

wb = load_workbook('test.xlsx')

print(wb.sheetnames)

