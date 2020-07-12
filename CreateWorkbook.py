from openpyxl import load_workbook

wb = load_workbook('test.xlsx')

sn = wb.sheetnames
print('Названия страниц -' 'sn')
print('_______________')
ws = wb.active
al = wb.active.title
print('Название активного листа -', al)
print('_______________')
sheet = wb['Data4export']
st = sheet.title
print(st)
print('_______________')
print('столбец А1 содержит', sheet['A1'].value)
print('столбец B1 содержит', sheet['b1'].value)
print('столбец C1 содержит', sheet['c1'].value)
print('столбец D1 содержит', sheet['d1'].value)
print('_______________')
sheet.cell(row=1, column=2).value
for i in range(1, 3):
     print(i, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value)
     print(i, sheet.cell(row=i, column=1).value, sheet.cell(row=i, column=2).value, sheet.cell(row=i, column=3).value, sheet.cell(row=i, column=4).value)
print('_______________')
for x in range(1, 2):
     for y in range(1, 4):
          ws.cell(row=x, column=y)
          print(ws.cell(row=x, column=y).value)
print('_______________')
cell_range = ws['A1':'D2']
print(cell_range)
ws4 = wb['RewritedData']
ws4['a1'] = sheet['d1'].value
ws4['b1'] = sheet['c1'].value
ws4['c1'] = sheet['b1'].value
ws4['d1'] = sheet['a1'].value
wb.save('test.xlsx')

