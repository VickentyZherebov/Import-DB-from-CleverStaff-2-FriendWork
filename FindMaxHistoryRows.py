from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

Workbook = load_workbook('AIHUB.xlsx')
history_sheet: Worksheet = Workbook['история']

print(history_sheet.title)


def fmgr(x, y, z):

    for x in range(3, history_sheet.max_row + 1):
        local_id = history_sheet[f'D{x}'].value

        if local_id is not None and y > z:
            y = y + 1

        if local_id is None:
            y = y + 1
            z = y

            return y, z


fmgr(None, 0, 0)

print(fmgr(None, 0, 0))
